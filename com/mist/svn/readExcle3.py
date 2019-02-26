from openpyxl import load_workbook
from copy import copy
import datetime


def write_data(logs,template_path,output_file_path):


    wb = load_workbook(template_path)
    wb.guess_types = True  # 猜测格式类型
    ws = wb.active
    logsCount=len(logs)
    # 以倒序的方式循环遍历第 N+1 行后的单元格，将其值复制到 N+M 行后的单元格中
    for i in range(ws.max_row, 6-1, -1):
        ws.row_dimensions[i+logsCount].height = copy(ws.row_dimensions[i].height)
        for j in range(1, 5):
            cell=ws.cell(row=i + logsCount, column=j)
            print(str(i)+':'+str(j))
            source=ws.cell(row=i , column=j)

            cell.value = source.value
            copyStyle(cell,source)
    print('Done!')
    # # 循环将第 N+1 行到第 N+M 行，将其值修改为空行
    for i in range ((5), 5+logsCount-1,+1):

        for j in range (1, 5):

            cell = ws.cell(row=i , column=j)

            source = ws.cell(row=5, column=j)
            cell.value = str(i - 5 + 1) + '、' + logs[i - 5]
            ws.row_dimensions[i].height = copy(ws.row_dimensions[5].height)
            copyStyle(cell, source)

    print('Done!')


            # ws.row_dimensions[i].height = copy(ws.row_dimensions[5].height)
            # cell.fill = copy(source.fill)
            # cell.border = copy(source.border)
            # cell.alignment = copy(source.alignment)
            # cell.number_format = copy(source.number_format)
            # cell.font = copy(source.font)

    endDate=datetime.datetime.now().strftime('%Y{}%m{}%d{}').format('年','月','日')
    startDate=(datetime.datetime.now()-datetime.timedelta(days=5)).strftime('%Y{}%m{}%d{}').format('年','月','日')
    ws.cell(2,1).value='周报: {}-{}'.format(startDate,endDate)

    wb.save(output_file_path)

# 复制单元格样式
def copyStyle(cell, source):
    cell.fill = copy(source.fill)
    cell.border = copy(source.border)
    cell.alignment = copy(source.alignment)
    cell.number_format = copy(source.number_format)
    cell.font = copy(source.font)